import os
import shutil

import environ
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

env = environ.Env()

ENGINE = create_engine(
    'postgresql+psycopg2://{}:{}@{}:{}/{}'.format(env("DB_USER"), env("DB_PASSWORD"),
                                                  env("DB_HOST"), env("DB_PORT"), env("DB_NAME")))
BASE_DIR = r'C:\Users\d_pio\PycharmProjects\air_pollution\pollution_data'


def create_directory(path):
    os.makedirs(path, exist_ok=True)


def move_file(file_path, dir_path_for_file):
    shutil.move(file_path, dir_path_for_file)


def update_table(df: pd.DataFrame, table_name: str):
    df.to_sql(
        table_name,
        con=ENGINE,
        if_exists="append",
        index=False,
        method="multi",
        chunksize=1000
    )


def file_converter(path: str) -> pd.DataFrame:
    workbook = load_workbook(path)
    ws = workbook.active

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    station_codes = first_row[1:]

    second_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    pollutant = second_row[1]

    first_col_values = [cell[0] for cell in ws.iter_rows(min_row=4, max_col=1, values_only=True)]
    measurement_columns = list(ws.iter_cols(min_col=2, min_row=4, values_only=True))

    rows = []
    for station_idx, station in enumerate(station_codes):
        column_data = measurement_columns[station_idx]
        for data, measurement in zip(first_col_values, column_data):
            if measurement is None or measurement < 0:
                continue
            rows.append({
                "pollutant": pollutant,
                "measurement_date": data,
                "station": station,
                "measurement_value": measurement
            })

    return pd.DataFrame(rows)


def process_file(file_name: str):
    file_path = os.path.join(BASE_DIR, file_name)
    year = file_name.split("_")[0]
    target_dir = os.path.join(BASE_DIR, f"{year}_pollutions_data")
    create_directory(target_dir)
    df_converted = file_converter(file_path)
    df_converted.columns = ['pollutant', 'measurement_date', 'station', 'measurement_value']
    update_table(df_converted, 'measurements')
    move_file(file_path, os.path.join(target_dir, file_name))
    print(f"Finished: {file_name}")


def main():
    for file in os.listdir(BASE_DIR):
        if file.endswith('.xlsx'):
            try:
                process_file(file)
            except Exception as e:
                print(f"[ERROR] File: {file}: {e}")


if __name__ == "__main__":
    main()
