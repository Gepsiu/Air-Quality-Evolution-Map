from pathlib import Path

import environ
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine

env = environ.Env()
PROJECT_PATH = Path(__file__).resolve().parent.parent

ENGINE = create_engine(
    'postgresql+psycopg2://{}:{}@{}:{}/{}'.format(env("DB_USER"), env("DB_PASSWORD"),
                                                  env("DB_HOST"), env("DB_PORT"), env("DB_NAME")))
FILE_PATH = f'{PROJECT_PATH}\Metadane oraz kody stacji i stanowisk pomiarowych.xlsx'


def update_table(df: pd.DataFrame, table_name: str):
    df.to_sql(
        table_name,
        con=ENGINE,
        if_exists="append",
        index=False,
        method="multi",
        chunksize=1000
    )


def file_converter(path: str) -> pd.DataFrame:
    workbook = load_workbook(path)
    ws = workbook.active
    regex = r"^\s*$"

    station_code = next(ws.iter_cols(min_col=2, min_row=2, max_col=2, values_only=True))
    voivodeship = next(ws.iter_cols(min_col=11, min_row=2, max_col=11, values_only=True))
    outdated_station_code = next(ws.iter_cols(min_col=5, min_row=2, max_col=5, values_only=True))
    df = pd.DataFrame({
        "code": station_code,
        "voivodeship": voivodeship
    }).dropna(how="all")
    df_outdated = pd.DataFrame({
        "code": outdated_station_code,
        "voivodeship": voivodeship
    }).dropna(how="all")
    df = df.replace(regex, pd.NA, regex=True)
    df = df.dropna(subset=["code"])
    df_outdated = df_outdated.replace(regex, pd.NA, regex=True)
    df_outdated = df_outdated.dropna(subset=["code"])
    df_outdated["code"] = df_outdated["code"].astype(str).str.split(",")
    df_outdated = df_outdated.explode("code")
    df_outdated["code"] = df_outdated["code"].str.strip()
    df_outdated = df_outdated.replace(regex, pd.NA, regex=True)
    df_outdated = df_outdated.dropna(subset=["code"])
    df = pd.concat([df, df_outdated], ignore_index=True)
    df = df.drop_duplicates(subset=["code"])
    return df


def process_file(path):
    df_converted = file_converter(path)
    df_converted.columns = ['code', 'voivodeship', 'outdated_code']
    update_table(df_converted, 'stations')
    file_name = Path(path).name
    print(f"Finished: {file_name}")


def main():
    try:
        process_file(FILE_PATH)
    except Exception as e:
        print(f"[ERROR] File: {FILE_PATH}: {e}")


if __name__ == "__main__":
    main()
